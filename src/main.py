import os
import json
import time
import hmac
import math
import random
import string
import hashlib
import binascii
import threading
import unicodedata

import pyaes
import requests
import openpyxl

from datetime import datetime, timedelta
from utils.utils_lib import LoggerManager, Utils

Utils.sync_work_dir()
logger = LoggerManager()


class AESCipher:
    def __init__(self, encryptionKey: bytes, initializationVector: bytes):
        self.encryption_key = encryptionKey
        self.initialization_vector = initializationVector

    def pad_text(self, text_bytes: bytes):
        """PKCS#7 填充"""
        padLength = 16 - len(text_bytes) % 16
        return text_bytes + bytes([padLength] * padLength)

    def encrypt_text(self, plain_text: str):
        textBytes = plain_text.encode('utf-8')
        paddedText = self.pad_text(textBytes)
        aes = pyaes.AESModeOfOperationCBC(
            key=self.encryption_key, iv=self.initialization_vector)
        cipherText = aes.encrypt(paddedText)
        return binascii.b2a_hex(cipherText).decode('utf-8').upper()


class Account:
    def __init__(self, username, account, password, task_advance_day, subject_filter, speed):
        self.username = username
        self.account = account
        self.password = password

        self.task_advance_day = task_advance_day
        self.subject_filter = subject_filter
        self.speed = speed
        self.lessons = []
        self.remaining_time = 0

        self.token = ''
        self.school_id = ''
        self.user_id = ''

        self.key = b'20171109124536982017110912453698'
        self.iv = b'2017110912453698'
        self.ip = f'{random.randint(59, 61)}.{random.randint(0, 230)}.{random.randint(0, 230)}.{random.randint(0, 230)}'
        self.common_header = {
            'content-type': 'text/plain',
            'access-control-allow-origin': '*',
            'origin': 'https://teacher.ewt360.com',
            'referer': 'https://teacher.ewt360.com/',
            'sec-ch-ua': '"Chromium";v="104", " Not A;Brand";v="99", "Microsoft Edge";v="104"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': 'Windows',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'token': '',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.30'
        }
        self.common_header_json = {
            'content-type': 'application/json',
            'access-control-allow-origin': '*',
            'origin': 'https://teacher.ewt360.com',
            'referer': 'https://teacher.ewt360.com/',
            'sec-ch-ua': '"Chromium";v="104", " Not A;Brand";v="99", "Microsoft Edge";v="104"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': 'Windows',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'token': '',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.30'
        }
        self.aes_cipher = AESCipher(self.key, self.iv)

    def get_now_ts(self):
        """获取当前时间戳"""
        return math.floor(time.time() * 1000)

    def get_midnight_ts(self, day: int):
        """获取午夜时间戳"""
        midnight = datetime.combine(datetime.today(), datetime.min.time())
        target_midnight = midnight + timedelta(days=day)
        return int(target_midnight.timestamp() * 1000)

    def ts_to_date(self, timestamp: int):
        """将时间戳转换为可阅读的日期格式"""
        return time.strftime('%Y-%m-%d', time.localtime(timestamp / 1000))

    def user_login(self):
        """登录用户账号, 刷新token"""
        ts = self.get_now_ts()
        try:
            res = requests.post(
                url='https://gateway.ewt360.com/api/authcenter/v2/oauth/login/account',
                data=json.dumps({
                    'autoLogin': 'true',
                    'password': self.aes_cipher.encrypt_text(self.password),
                    'platform': 1,
                    'userName': self.account
                }),
                headers={
                    'accept': 'application/json',
                    'accept-language': 'zh-CN,zh;q=0.9',
                    'content-type': 'application/json;charset=UTF-8',
                    'origin': 'https://web.ewt360.com',
                    'platform': '1',
                    'referer': 'https://web.ewt360.com/',
                    'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="100", "Google Chrome";v="100"',
                    'sec-ch-ua-mobile': '?0',
                    'sec-ch-ua-platform': '"Windows"',
                    'sec-fetch-dest': 'empty',
                    'sec-fetch-mode': 'cors',
                    'sec-fetch-site': 'same-site',
                    'secretid': '2',
                    'sign': hashlib.md5(str(ts).encode('utf8') + 'bdc739ff2dcf'.encode('utf8')).hexdigest().upper(),
                    'timestamp': str(ts),
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.30'
                }
            )

            data = res.json()
            if data.get('code', 500) == '200':
                self.token = data['data']['token']
                self.common_header['token'] = self.token
                self.common_header_json['token'] = self.token
            else:
                e_msg = f'[{self.username}] 登录状态码异常: {data}'
                logger.error(e_msg)
                raise Exception(e_msg)

        except Exception as e:
            e_msg = f'[{self.username}] 登录请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def get_user_info(self):
        """获取用户信息"""
        try:
            res = requests.get(
                url='https://teacher.ewt360.com/api/eteacherproduct/school/getSchoolUserInfo',
                headers=self.common_header
            )

            data = res.json()
            if data.get('code', 500) == '200':
                self.school_id = data['data']['schoolId']
                self.user_id = data['data']['userId']
            else:
                e_msg = f'[{self.username}] 获取用户信息失败，状态码异常：{data}'
                logger.error(e_msg)
                raise Exception(e_msg)

        except Exception as e:
            e_msg = f'[{self.username}] 获取用户信息请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def get_scene_id(self):
        """获取场景ID"""
        try:
            res = requests.get(
                url=(
                    'https://gateway.ewt360.com/api/holidayprod/scene/student/study/checkHoliday?'
                    f'clientType=1&preview=0&schoolId={self.school_id}&timestamp={self.get_now_ts()}'
                ),
                headers=self.common_header
            )

            data = res.json()
            if data.get('code', 500) == '200':
                return data['data']['sceneList']
            else:
                e_msg = f'[{self.username}] 获取场景ID失败，状态码异常：{data}'
                logger.error(e_msg)
                raise Exception(e_msg)

        except Exception as e:
            e_msg = f'[{self.username}] 获取场景ID请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def get_homework_id(self, scene_id: int):
        """获取作业ID"""
        try:
            res = requests.get(
                url=(
                    'https://gateway.ewt360.com/api/homeworkprod/homework/student/holiday/'
                    f'getHomeworkSummaryInfo?schoolId={self.school_id}&timestamp={self.get_now_ts()}&sceneId={scene_id}'
                ),
                headers=self.common_header
            )

            data = res.json()
            if data.get('code', 500) == '200':
                return data['data']['homeworkIds']
            else:
                e_msg = f'[{self.username}] 获取作业ID失败，状态码异常：{data}'
                logger.error(e_msg)
                raise Exception(e_msg)

        except Exception as e:
            e_msg = f'[{self.username}] 获取作业ID请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def get_day_list(self, homeworkId: int, sceneId: int):
        """获取天数表"""
        try:
            res = requests.post(
                url=(
                    'https://gateway.ewt360.com/api/homeworkprod/homework/student/holiday/'
                    f'getHomeworkDistribution?sceneId={sceneId}'
                ),
                data=json.dumps({
                    'homeworkIds': [homeworkId],
                    'type': 2,
                    'isSelfTask': 'false',
                    'userOptionTaskId': 'null',
                    'schoolId': self.school_id,
                    'sceneId': sceneId
                }),
                headers=self.common_header_json
            )

            data = res.json()
            if data.get('code', 500) == '200':
                return data['data']['days']
            else:
                e_msg = f'[{self.username}] 获取天数表失败，状态码异常：{data}'
                logger.error(e_msg)
                raise Exception(e_msg)

        except Exception as e:
            e_msg = f'[{self.username}] 获取天数表请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def get_homework_list(self, homeworkId: int, day: list, sceneId: int):
        """获取作业列表"""
        try:
            res = requests.post(
                url=(
                    'https://gateway.ewt360.com/api/homeworkprod/homework/student/holiday/'
                    f'pageHomeworkTasks?sceneId={sceneId}'
                ),
                data=json.dumps({
                    'dayId': day['dayId'][0],
                    'day': day['day'][0],
                    'status': 0,
                    'homeworkIds': [homeworkId],
                    'isSelfTask': 'false',
                    'userOptionTaskId': 'null',
                    'pageIndex': 1,
                    'pageSize': 30,
                    'missionType': 0,
                    'schoolId': self.school_id,
                    'sceneId': sceneId
                }),
                headers=self.common_header_json
            )

            data = res.json()
            if data.get('code', 500) == '200':
                return data['data']['data']
            else:
                e_msg = f'[{self.username}] 获取作业列表失败，状态码异常：{data}'
                logger.error(e_msg)
                raise Exception(e_msg)

        except Exception as e:
            e_msg = f'[{self.username}] 获取作业列表请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def get_player_global_conf(self):
        """获取视频播放器全局配置"""
        try:
            res = requests.get(
                url=(
                    'https://web.ewt360.com/api/videoplayerprod/videoplayer/getPlayerGlobalConf?'
                    f'videoBizCode=1001&sdkVersion=3.0.19&_={self.get_now_ts()}'
                ),
                headers={
                    'Token': self.token,
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.30'
                }
            )

            data = res.json()
            if data.get('code', 500) == '200':
                secret = data['data']['globalInfo']['secret']
                x_bfe_session_id = data['data']['globalInfo']['sessionId']
                begin_ts = data['data']['globalInfo']['ts']
                return secret, x_bfe_session_id, begin_ts
            else:
                e_msg = f'[{self.username}] 获取视频播放器全局配置信息失败，状态码异常：{data}'
                logger.error(e_msg)
                raise Exception(e_msg)

        except Exception as e:
            e_msg = f'[{self.username}] 获取视频播放器全局配置信息请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def upload_progress(self, secret, x_bfe_session_id, begin_time, report_time, lesson_id, course_id, index, action, duration):
        """上传进度"""

        def generate_uuid(index):
            """生成UUID"""
            return ''.join(random.sample(string.ascii_letters + string.digits, 8)) + f'_{index}'

        def generate_signature(secret, action, duration, report_time, token):
            """生成签名"""
            signature_str = f'action={action}&duration={duration}&mstid={token}&signatureMethod=HMAC-SHA1&signatureVersion=1.0&timestamp={report_time}&version=2022-08-02'
            return hmac.new(bytes(secret, encoding='utf-8'), signature_str.encode('utf-8'), hashlib.sha1).hexdigest()

        uuid = generate_uuid(index)
        ts = self.get_now_ts()
        url = (
            f"https://bfe.ewt360.com/monitor/web/collect/batch?TrVideoBizCode=1013&TrFallback=0&TrUserId={self.user_id}"
            f"&TrlessonId={lesson_id}&TrUuId={uuid}&sdkVersion=3.0.8&_={ts}"
        )
        headers = {
            'access-control-allow-origin': '*',
            'content-type': 'application/json',
            'origin': 'https://web.ewt360.com',
            'referer': 'https://web.ewt360.com/',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-site',
            'Token': self.token,
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.30',
            'X-Bfe-Session-ID': x_bfe_session_id
        }
        signature = generate_signature(
            secret, action, duration, report_time, self.token)
        payload = {
            'CommonPackage': {
                'userid': int(self.user_id),
                'ip': self.ip,
                'os': 'Windows',
                'resolution': '1920*1080',
                'mstid': self.token,
                'browser': 'Chrome',
                'browser_ver': '5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.30',
                'playerType': 1,
                'sdkVersion': '3.0.19',
                'videoBizCode': '1013',
            },
            'EventPackage': [{
                'lesson_id': str(lesson_id),
                'course_id': str(course_id),
                'stay_time': duration,
                'status': 3 if action != 2 else 1,
                'begin_time': str(begin_time),
                'report_time': report_time,
                'point_time_id': 1,
                'point_time': 60000,
                'point_num': 25,
                'video_type': 1,
                'speed': self.speed,
                'quality': '高清',
                'action': action,
                'fallback': 0,
                'uuid': uuid
            }],
            'signature': signature,
            'sn': '',
            '_': ts
        }

        try:
            res = requests.post(url, data=json.dumps(payload), headers=headers)
            data = res.json()
            if data.get('code', 500) == 200:
                return None
            else:
                e_msg = f'[{self.username}] 上传进度失败，状态码异常：{data}'
                logger.error(e_msg)
                raise Exception(e_msg)
        except Exception as e:
            e_msg = f'[{self.username}] 上传进度请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def upload_fm_progress(self, cls: list):
        """上传FM进度"""
        try:
            res = requests.post(
                url=f"https://gateway.ewt360.com/api/homeworkprod/homework/student/updateMission?sceneId={cls['sceneId']}",
                data=json.dumps({
                    'contentId': cls['contentId'],
                    'contentType': 3,
                    'sceneId': cls['sceneId'],
                    'percent': 1,
                    'schoolId': self.school_id
                }),
                headers=self.common_header_json
            )

            data = res.json()
            if data.get('code', 500) == '200':
                return None
            else:
                e_msg = f'[{self.username}] 上传FM进度失败，状态码异常：{data}'
                logger.error(e_msg)
                raise Exception(e_msg)

        except Exception as e:
            e_msg = f'[{self.username}] 上传FM进度请求失败: {e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def is_lesson_completed(self, cls: list):
        """检查课程是否已完成"""
        network_list = self.get_homework_list(
            cls['homeworkId'], {'day': cls['day'], 'dayId': cls['dayId']}, cls['sceneId'])

        for item in network_list:
            if item['contentId'] == cls['contentId'] and item['parentContentId'] == cls['parentContentId']:
                return False

        return True

    def update_progress_and_output_time(self, cls: list):
        """更新视频进度并输出剩余时间"""

        def update_local_data(ratio, cls):
            """更新本地数据"""
            self.remaining_time = 0
            for lesson in self.lessons:
                if (lesson['contentId'] == item['contentId'] and
                    lesson['parentContentId'] == item['parentContentId'] and
                        lesson['day'] == item['day']):
                    lesson['ratio'] = ratio
            if item['ratio'] < 1 and item['contentType'] != 3:
                self.remaining_time += (item['ratio'] - 0.8) * \
                    item['duration'] / self.speed + 5

            self.remaining_time = round(self.remaining_time / 60, 2)
            logger.info(f'[{self.username}] 总剩余时间: {self.remaining_time} Min')

        network_list = self.get_homework_list(
            cls['homeworkId'], {'day': cls['day'], 'dayId': cls['dayId']}, cls['sceneId'])
        for item in network_list:
            if item['contentId'] == cls['contentId'] and item['parentContentId'] == cls['parentContentId']:
                update_local_data(item['ratio'], cls)
                return
        update_local_data(1, cls)

    def log_lesson_details(self, cls: list | None = None,
                           day_show_width: int = 15, subject_name_width: int = 15,
                           title_width: int = 50, ratio_width: int = 10):
        """日志输出课程详细信息"""

        def get_display_width(text):
            """获取显示宽度"""
            return sum(2 if unicodedata.east_asian_width(char) in ('F', 'W') else 1 for char in text)

        def pad_text(text, totalWidth):
            padding = totalWidth - get_display_width(text)
            return text + ' ' * max(padding, 0)

        def format_lesson_details(lesson, day_show_width: int = 15, subject_name_width: int = 15, title_width: int = 50, ratio_width: int = 10):
            """格式化课程详细信息"""
            day_show = pad_text(lesson['dayShow'], day_show_width)
            subject_name = pad_text(lesson['subjectName'], subject_name_width)
            title = pad_text(lesson['title'], title_width)
            ratio = pad_text(
                f"{round(lesson['ratio'] * 100 / 0.8, 2)}%", ratio_width)
            return f'{day_show} {subject_name} {title} {ratio}'

        if cls is None:
            logger.info(f'[{self.username}]')
        else:
            logger.info(f'[{self.username}] 课程表：')

        logger.info(
            f'{pad_text('日期', day_show_width)} {pad_text('学科', subject_name_width)} ' +
            f'{pad_text('标题', title_width)} {pad_text('进度', ratio_width)}'
        )

        lessons = self.lessons if cls is None else [cls]
        for lesson in lessons:
            logger.info(
                format_lesson_details(
                    lesson, day_show_width, subject_name_width, title_width, ratio_width)
            )

    def scan_lessons_and_run_filter(self):
        """扫描课程并运行课程过滤器"""
        midnight_ts = self.get_midnight_ts(self.task_advance_day)
        try:
            for scene in self.get_scene_id():
                scene_id = scene['id']
                for homework_id in self.get_homework_id(scene_id):
                    day_list = self.get_day_list(homework_id, scene_id)
                    valid_days = [
                        day for day in day_list if day['day'][0] <= midnight_ts]
                    for day in valid_days:
                        for cls in self.get_homework_list(homework_id, day, scene_id):
                            if cls['subjectId'] in self.subject_filter:
                                continue
                            if cls['contentType'] == 2:
                                continue
                            self.lessons.append({
                                'subjectId': cls['subjectId'],
                                'parentContentId': cls['parentContentId'],
                                'title': cls['title'],
                                'duration': cls['duration'],
                                'ratio': cls['ratio'],
                                'subjectName': cls['subjectName'],
                                'contentId': cls['contentId'],
                                'contentType': cls['contentType'],
                                'homeworkId': homework_id,
                                'sceneId': scene_id,
                                'day': day['day'][0],
                                'dayId': day.get('dayId', [0])[0],
                                'dayShow': self.ts_to_date(day['day'][0])
                            })
            if len(self.lessons) == 0:
                logger.info(f'[{self.username}] 任务已全部完成')
                raise SystemExit

        except Exception as e:
            e_msg = f'[{self.username}] 课程获取失败：{e}'
            logger.error(e_msg)
            raise Exception(e_msg)

    def run_lesson(self, cls: list):
        """以课程为单位刷课"""
        lesson_id = cls['contentId']
        course_id = cls['parentContentId']
        lesson_duration = cls['duration']
        logger.info(f'[{self.username}] 课程任务已开始：{cls['title']}')

        if cls['contentType'] == 3:
            self.upload_fm_progress(cls)
            return

        time_finished_ms = 0
        residue_time_ms = round((lesson_duration * 1000 * 0.8 -
                                lesson_duration * 1000 * cls['ratio'] + 5000) / self.speed)
        secret, x_bfe_session_id, begin_ts = self.get_player_global_conf()
        begin_time = int(begin_ts)
        logger.info(f'[{self.username}] 获取认证信息成功')

        self.upload_progress(secret, x_bfe_session_id, begin_time,
                             self.get_now_ts(), lesson_id, course_id, 1, 1, 0)
        logger.info(f'[{self.username}] 初始进度上传成功')

        segment_count = 2
        while time_finished_ms < residue_time_ms:
            logger.info(
                f'[{self.username}] 课程[{cls['title']}]还需{(residue_time_ms - time_finished_ms) / 1000}s可完成')

            duration = 60000 if (
                residue_time_ms - time_finished_ms) >= 60000 else 0
            if duration == 0 and (residue_time_ms - time_finished_ms) < 60000:
                if residue_time_ms - time_finished_ms < 0:
                    e_msg = f'课程运行异常，剩余时间<0：{residue_time_ms - time_finished_ms}'
                    logger.error(e_msg)
                    raise Exception(e_msg)
                duration = residue_time_ms - time_finished_ms

            time.sleep(duration / 1000)
            self.upload_progress(
                secret, x_bfe_session_id, begin_time, self.get_now_ts(),
                lesson_id, course_id, segment_count, 2, duration
            )
            time_finished_ms += duration
            segment_count += 1
            self.update_progress_and_output_time(cls)
            self.log_lesson_details(cls=cls)

        self.upload_progress(
            secret, x_bfe_session_id, begin_time, self.get_now_ts(),
            lesson_id, course_id, segment_count, 3, 0
        )

        if self.is_lesson_completed(cls):
            logger.info(f'[{self.username}] 课程已完成：{cls['title']}')

    def run_account(self):
        """以账号为单位刷课"""
        try:
            logger.info(
                f'[{self.username}] 账号：{self.account}, 密码：{self.password}, 速度：{self.speed}, '
                f'计划提前完成天数：{self.task_advance_day}, 科目过滤：{self.subject_filter}'
            )

            self.user_login()
            logger.info(f'[{self.username}] 登录成功')

            self.get_user_info()
            logger.info(f'[{self.username}] 获取用户信息成功，课程扫描中...')

            self.scan_lessons_and_run_filter()
            logger.info(f'[{self.username}] 开始运行任务，总数：{len(self.lessons)}个')

            self.update_progress_and_output_time(self.lessons[0])
            self.log_lesson_details()

            for lesson in self.lessons:
                self.run_lesson(lesson)

            logger.info(f'[{self.username}] 任务已全部完成')

        except Exception as e:
            e_msg = f'[{self.username}] 任务线程异常退出：{e}'
            logger.error(e_msg)
            return


if __name__ == '__main__':
    # 多账号模式配置
    multi_account = True

    excel_path = 'accounts.xlsx'
    max_accounts, account_run_freq = (1000, 0)
    subject_filter = []
    speed = 2
    task_advance_day = 30

    if not multi_account:
        username, account, password = ('test', '', '985211')
        Account(
            username, account, password, 
            task_advance_day, subject_filter, speed
        ).run_account()
    else:
        threads = []
        wb = openpyxl.load_workbook(os.path.join(Utils.get_bundle_dir(), excel_path))
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        column_mapping = {
            '姓名': header_row.index('姓名'),
            'EWT账号': header_row.index('EWT账号'),
            '密码': header_row.index('密码')
        }
        row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row_count >= max_accounts:
                break
            username = row[column_mapping['姓名']]
            account = str(row[column_mapping['EWT账号']])
            password = str(row[column_mapping['密码']])
            thread = threading.Thread(
                target=Account(
                    username, account, password, 
                    task_advance_day, subject_filter, speed
                ).run_account
            )
            thread.start()
            threads.append(thread)
            row_count += 1
            time.sleep(account_run_freq)

        wb.close()
        time.sleep(account_run_freq + 20)
        logger.info(
            f'总循环数量：{row_count}，实际运行中的账号数量：{sum(1 for t in threads if t.is_alive())}')
        
        for thread in threads:
            thread.join()
